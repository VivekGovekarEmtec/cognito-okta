import base64
import io
from openpyxl import load_workbook
from common_component.src.core.repositories.data_repository import call_postgres_function, CreateDBConnection
from common_component.src.core.utils.Logger import Log
from common_component.src.core.helpers.encoder import jsonable_encoder

log = Log().get_logger_service("Get station information")
# Create an instance of the CreateDBConnection class
db_instance = CreateDBConnection()


# price change auth
def process_excel_file(base64_data):
    """
    This API is used to process data from an Excel file.
    """
    try:
        log.append_keys(service_function="process_excel_file")
        log.debug("Entered into process_excel_file service")
        # Decode the base64 string
        binary_data = base64.b64decode(base64_data)
        excel_io = io.BytesIO(binary_data)
        # Load the Excel workbook
        workbook = load_workbook(excel_io)
        sheet = workbook['Schedule A']

        def get_expected_value(row, column):
            if row == 13 and column == 3:
                return "base wholesale price"
            elif row == 13 and column == 4:
                return "fed. excise tax"
            elif row == 13 and column == 5:
                return "prov. tax"
            elif row == 13 and column == 6:
                return "wholesale selling price"
            elif row == 13 and (column == 7 or column == 9):
                return "min"
            elif row == 13 and (column == 8 or column == 10):
                return "max"
            elif row in [14, 19, 24, 29, 34, 39] and column == 2:
                return f"zone {((row - 14) // 5) + 1}"
            else:
                return None

        for row in [13, 14, 19, 24, 29, 34, 39]:
            for column in range(3, 11) if row == 13 else [2]:
                cell_value = sheet.cell(row=row, column=column).value
                cell_lowered = cell_value.strip().lower() if cell_value is not None else None
                get_expected_value(row, column)
                if cell_lowered is None or cell_lowered != get_expected_value(row, column):
                    raise ValueError("INVALIDSTRUCTURE")

        zone_data = {
            "zone1": {
                "regular_min": sheet.cell(row=15, column=7).value,
                "regular_max": sheet.cell(row=15, column=8).value,
                "plus_min": sheet.cell(row=16, column=7).value,
                "plus_max": sheet.cell(row=16, column=8).value,
                "supreme_min": sheet.cell(row=17, column=7).value,
                "supreme_max": sheet.cell(row=17, column=8).value,
                "diesel_min": sheet.cell(row=18, column=7).value,
                "diesel_max": sheet.cell(row=18, column=8).value,
            },
            "zone2": {
                "regular_min": sheet.cell(row=20, column=7).value,
                "regular_max": sheet.cell(row=20, column=8).value,
                "plus_min": sheet.cell(row=21, column=7).value,
                "plus_max": sheet.cell(row=21, column=8).value,
                "supreme_min": sheet.cell(row=22, column=7).value,
                "supreme_max": sheet.cell(row=22, column=8).value,
                "diesel_min": sheet.cell(row=23, column=7).value,
                "diesel_max": sheet.cell(row=23, column=8).value,
            },
            "zone3": {
                "regular_min": sheet.cell(row=25, column=7).value,
                "regular_max": sheet.cell(row=25, column=8).value,
                "plus_min": sheet.cell(row=26, column=7).value,
                "plus_max": sheet.cell(row=26, column=8).value,
                "supreme_min": sheet.cell(row=27, column=7).value,
                "supreme_max": sheet.cell(row=27, column=8).value,
                "diesel_min": sheet.cell(row=28, column=7).value,
                "diesel_max": sheet.cell(row=28, column=8).value,
            },
            "zone4": {
                "regular_min": sheet.cell(row=30, column=7).value,
                "regular_max": sheet.cell(row=30, column=8).value,
                "plus_min": sheet.cell(row=31, column=7).value,
                "plus_max": sheet.cell(row=31, column=8).value,
                "supreme_min": sheet.cell(row=32, column=7).value,
                "supreme_max": sheet.cell(row=32, column=8).value,
                "diesel_min": sheet.cell(row=33, column=7).value,
                "diesel_max": sheet.cell(row=33, column=8).value,
            },
            "zone5": {
                "regular_min": sheet.cell(row=35, column=7).value,
                "regular_max": sheet.cell(row=35, column=8).value,
                "plus_min": sheet.cell(row=36, column=7).value,
                "plus_max": sheet.cell(row=36, column=8).value,
                "supreme_min": sheet.cell(row=37, column=7).value,
                "supreme_max": sheet.cell(row=37, column=8).value,
                "diesel_min": sheet.cell(row=38, column=7).value,
                "diesel_max": sheet.cell(row=38, column=8).value,
            },
            "zone6": {
                "regular_min": sheet.cell(row=40, column=7).value,
                "regular_max": sheet.cell(row=40, column=8).value,
                "plus_min": sheet.cell(row=41, column=7).value,
                "plus_max": sheet.cell(row=41, column=8).value,
                "supreme_min": sheet.cell(row=42, column=7).value,
                "supreme_max": sheet.cell(row=42, column=8).value,
                "diesel_min": sheet.cell(row=43, column=7).value,
                "diesel_max": sheet.cell(row=43, column=8).value,
            },
            "is_read_only": True
        }
        response = {
            "status_code": 200,
            "message": "Excel file processing successful",
            "data": {"zone_data": zone_data}
        }
    except ValueError as ve:
        raise ve
    except Exception as exc:
        raise exc
    else:
        log.debug("process_excel_file service executed successfully")
        return jsonable_encoder(response)
