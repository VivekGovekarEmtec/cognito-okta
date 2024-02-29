from sqlalchemy import text
import base64
import io
from openpyxl import load_workbook
from common_component.src.core.repositories.data_repository import call_postgres_function, CreateDBConnection
from common_component.src.core.utils.Logger import Log
from common_component.src.core.helpers.encoder import jsonable_encoder

log = Log().get_logger_service("Get station information")
# Create an instance of the CreateDBConnection class
db_instance = CreateDBConnection()


# price change auth
def is_regulated_prices_file_exists(file_name, base64_data):
    """
    This API is used to check if regulated prices header file exists
    """
    try:
        log.append_keys(service_function="is_regulated_prices_file_exists")
        log.debug("Entered into is_regulated_prices_file_exists service")
        # Decode the base64 string
        binary_data = base64.b64decode(base64_data)

        # Create a BytesIO object from the binary data
        excel_io = io.BytesIO(binary_data)
        workbook = load_workbook(excel_io, read_only=True)
        with db_instance.create_writer_connection() as db:

            if "Schedule A" not in workbook.sheetnames:
                raise ValueError("The <Schedule A> sheet name missing in the file")
            parameters = {"file_name": file_name}
            is_regulated_prices_header_file_exists_function_name = 'dbo.fn_regulated_prices_header_file_exist'
            query = text(
                "select * from " + is_regulated_prices_header_file_exists_function_name + "(CAST(:file_name AS "
                                                                                          "character varying))")

            output = call_postgres_function(query=query, parameters=parameters, db=db)
        log.debug("Received response from database")
        response = {
            "status_code": 200,
            "message": "is regulated prices header file exists response",
            "data": output[0]
        }
    except ValueError as ve:
        raise ve
    except Exception as exc:
        raise exc
    else:
        log.debug("is_regulated_prices_file_exists service executed successfully")
        return jsonable_encoder(response)
